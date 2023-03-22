import polars
import pandas
import xlwings
from openpyxl import load_workbook
# Show all sheets name in workbooks

def show_sheetnames(path) -> list:
    workbook = load_workbook(path)
    sheetnames = workbook.sheetnames
    return sheetnames

# def read_excel(path):
    


